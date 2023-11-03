import pandas as pd
import openpyxl
import numpy as np
import csv

sentances  = ["i want to make coffee:DA"]
sentances.append("where am i:LR")
sentances.append("what can i do here:AR")
sentances.append("prepare meal:DA")
sentances.append("use refrigerator:DA")
sentances.append("use fan:DA")
sentances.append("use oven:DA")
sentances.append("use stove:DA")
sentances.append("i would like to wash sheets:DA")
sentances.append("i would like to watch television:DA")
sentances.append("wash sheets:DA")
sentances.append("wash clothes:DA")
sentances.append("cook dinner:DA")
sentances.append("i would like to cook dinner:DA")
sentances.append("where is the coffee:DA")
sentances.append("rewind video:RV")
sentances.append("forward:FF")

sentances.append("pause video:PV")
sentances.append("skip:FF")
sentances.append("rewind:RV")
sentances.append("pause:PV")
sentances.append("fast forward:FF")
sentances.append("play video:SV")
sentances.append("resume:SV")
sentances.append("resume video:SV")
sentances.append("negative:SN")
sentances.append("positive:SN")
sentances.append("yes:SN")
sentances.append("no:SN")
sentances.append("yes:SN")
sentances.append("no:SN")
sentances.append("shutdown app:SD")
sentances.append("shut:SD")
sentances.append("hello:NA")
#sentances.append("play video")
sentances.append("shut down app:SD")
sentances.append("shut:SD")
#sentances.append("play video")
sentances.append("negative:SN")
sentances.append("positive:SN")
sentances.append("correct:SN")
sentances.append("incorrect:SN")
sentances.append("fake:SN")
sentances.append("false:SN")
sentances.append("nay:SN")
sentances.append("wrong:SN")
sentances.append("yay:SN")

sentances.append("genuine:SN")
sentances.append("true:SN")
sentances.append("right:SN")
sentances.append("correct:SN")
sentances.append("incorrect:SN")
sentances.append("i want to make coffee:DA")
sentances.append("where am i:LR")
sentances.append("start camera:CC")
sentances.append("stop camera:CC")
sentances.append("open camera:CC")
sentances.append("close camera:CC")
sentances.append("start camera:CC")
sentances.append("disengage:CC")
sentances.append("display:CC")

sentances.append("start camera:CC")
sentances.append("stop camera:CC")
sentances.append("open camera:CC")
sentances.append("close camera:CC")

sentances.append("i would like to wash sheets:DA")
sentances.append("i would like to watch television:DA")
sentances.append("wash sheets:DA")
sentances.append("wash clothes:DA")
sentances.append("cook dinner:DA")
sentances.append("i would like to cook dinner:DA")
sentances.append("where is the coffee:DA")
sentances.append("rewind video:RV")
sentances.append("pause video:PV")
sentances.append("skip forward:FF")
sentances.append("shut down:SD")
sentances.append("turn off:SD")
sentances.append("return home:SD")
sentances.append("shutdown:SD")
sentances.append("power off:SD")
sentances.append("shut down system:SD")
sentances.append("system disable:SD")
sentances.append("disengage:SD")
sentances.append("what can I do here:AR")
sentances.append("activites:AR")
sentances.append("what activites are available:AR")
sentances.append("what activites are available at this location:AR")
sentances.append("what can I interactive with:AR")
sentances.append("what can I make:AR")
sentances.append("is there water here:AR")
sentances.append("what is here:AR")
sentances.append("where am I:LR")
sentances.append("where am I currently at:LR")
sentances.append("what is here:AR")
sentances.append("where am is the kitchen:NAV")
sentances.append("where am I currently at:LR")
sentances.append("am I currently in the kitchen:LR")
sentances.append("am I currently in the bedroom:LR")
sentances.append("am I currently in the bathroom:LR")
sentances.append("where am I at:LR")
sentances.append("can you tell me where I am:LR")
sentances.append("i need to know my current position:LR")
sentances.append("location:LR")
sentances.append("position:LR")
sentances.append("where:LR")
sentances.append("activities:AR")
sentances.append("do here:AR")
sentances.append("available:AR")
sentances.append("coffee:DA")
sentances.append("cook:DA")
sentances.append("diswasher:DA")
#sentances.append("society excited by cottage private an it esteems:NA")
#sentances.append("fully begin on by wound an:NA")
#sentances.append("girl rich in do up or both:NA")
#sentances.append("at declared in as rejoiced of together:NA")
#sentances.append("he impression collecting delightful unpleasant by prosperous as on:NA")
#sentances.append("end too talent she object mrs wanted remove giving:NA")
#sentances.append("the text is a corrupted version of the original and therefore does not mean anything in particular:NA")
#sentances.append("It was a scrape that he hardly noticed. Sure, there was a bit of blood but it was minor compared to most of the other cuts and bruises he acquired on his adventures. There was no way he could know that the rock that produced the cut had alien genetic material on it that was now racing through his bloodstream. He felt perfectly normal and continued his adventure with no knowledge of what was about to happen to him.:NA")
#sentances.append("Joe made the sugar cookies Susan decorated them.:NA")
#sentances.append("Being unacquainted with the chief raccoon was harming his prospects for promotion.:NA")
#sentances.append("She was sad to hear that fireflies are facing extinction due to artificial light, habitat loss, and pesticides.:NA")
sentances.append("go to the living room:NAV")
sentances.append("go to kitchen:NAV")
sentances.append("can I go to the living room:NAV")
sentances.append("take me to the sunroom:NAV")
sentances.append("kitchen to bathroom:NAV")
sentances.append("move to garage:NAV")
sentances.append("i want to go to the kitchen:NAV")
sentances.append("how do i get to the bathroom:NAV")
sentances.append("where is the bathroom:NAV")
sentances.append("i am lost take me to the livingroom:NAV")
sentances.append("help me find the kitchen:NAV")
sentances.append("help me to the bedroom:NAV")
sentances.append("what activites are available here:AR")
sentances.append("what activity’s can I do here:AR")
sentances.append("Am i able to complete a task here:AR")
sentances.append("What activities are here:AR")
sentances.append("What options do I have:AR")
sentances.append("What’s happening here:AR")
sentances.append("What are some activities that you offer:AR")
sentances.append("where am I at:LR")
sentances.append("can I eat in the kitchen:LR")
sentances.append("I don't know where I am, can you help me:LR")
sentances.append("What’s my current address :LR")
sentances.append("Where are you:LR")
sentances.append("Location please:LR")
sentances.append("Do you know what my location is:LR")
sentances.append("open the camera:CC")
sentances.append("where am I at:CC")
sentances.append("Open the camera for me:CC")
sentances.append("Open camera app:CC")
sentances.append("Open the camera:CC")
sentances.append("Bring up the camera:CC")
sentances.append("Activate the camera:CC")
sentances.append("Turn on camera:CC")
sentances.append("Take a picture :CC")
sentances.append("turn the camera off:CC")
sentances.append("Close the camera for me:CC")
sentances.append("Close camera app:CC")
sentances.append("Shut down the camera:CC")
sentances.append("Turn off the camera:CC")
sentances.append("Turn off camera:CC")
sentances.append("can i return to home:SD")
sentances.append("Close this app:SD")
sentances.append("Return to home page :SD")
sentances.append("Good night:SD")
sentances.append("Thank you :SD")
sentances.append("Exit:SD")
sentances.append("Turn off assistant:SD")
sentances.append("Deactivate assistance:SD")
sentances.append("Go home:SD")
sentances.append(" Return home:SD")
sentances.append("Go to the home page:SD")
sentances.append("pause the video:PV")
sentances.append("Stop the video:PV")
sentances.append("Pause:PV")
sentances.append("Freeze video:PV")
sentances.append("Stop video:PV")
sentances.append("Freeze video:PV")
sentances.append("Press pause:PV")
sentances.append("can you start the video:SV")
sentances.append("Resume the video:SV")
sentances.append("Play:SV")
sentances.append("Play video:SV")
sentances.append("Continue video:SV")
sentances.append("Unfreeze video:SV")
sentances.append("Replay video:SV")
sentances.append("Play video :SV")
sentances.append("Press play :SV")
sentances.append("play it back:RV")
sentances.append("Go back a few seconds:RV")
sentances.append("Turn back video:RV")
sentances.append("Reverse video:RV")
sentances.append("Restart video:RV")
sentances.append("Go back on the video:RV")
sentances.append("Go back:RV")
sentances.append("Minus 15 seconds :RV")
sentances.append("skip forward in the video:FF")
sentances.append("Next video:FF")
sentances.append("Fast forward :FF")
sentances.append("Skip video:FF")
sentances.append("Pass video:FF")
sentances.append("Speed up video:FF")
sentances.append("Skip:FF")
sentances.append("Skip forward 30 seconds :FF")
sentances.append("Go forward:FF")
sentances.append("i need help making a sandwich:DA")
sentances.append("How do I use the microwave?:DA")
sentances.append("How do I make coffee :DA")
sentances.append("How do I cook an egg:DA")
sentances.append("How do I do this:DA")
sentances.append("i need to go to the livingroom:NAV")
sentances.append("Take me to the garden:NAV")
sentances.append("Lead me to the bathroom :NAV")
sentances.append("Locate living room :NAV")
sentances.append("Take me to the living room:NAV")



df = pd.read_excel("./Keyword Map NLP.xlsx")
Loc = df.loc[df['Loc']==1]
#print(Loc)
workbook = openpyxl.load_workbook('./Keyword Map NLP.xlsx')
sh = workbook.active
'''                 
get= []
print("start")
for i in range(1,sh.max_column+1):
    #print(i)
    search = sh.cell(row=1,column=i).value
    #print(search)
    if(search == 'Loc'):
        for j in range(2,sh.max_row+1):
            one = sh.cell(row=j,column=i).value
            if(one == 1):
                #print(sh.cell(row=j,column=1).value)
                get.append(sh.cell(row=j,column=1).value)
        break
                 
for w in get:
    filer = (f"i want to go to the {w}:NAV")
    sentances.append(filer)

for w in get:
    filer = (f"navigate to {w}:NAV")
    sentances.append(filer)
for w in get:
    filer = (f"take me to {w}:NAV")
    sentances.append(filer)
for w in get:
    filer = (f"where is to {w}:NAV")
    sentances.append(filer)
for w in get:
    filer = (f"can i go to {w}:NAV")
    sentances.append(filer)


for w in get:
    filer = (f"is this the {w}:LR")
    sentances.append(filer)

for w in get:
    filer = (f"am I in the {w}:LR")
    sentances.append(filer)
'''

fields  = ["description","category"]
with open('text_train.csv', 'w') as csvfile:
    csvwriter = csv.writer(csvfile)  
    csvwriter.writerow(fields) 
    rows = []
    for data in sentances:
    # creating a csv writer object 
        
        
    # writing the fields 

        rows.append(data.split(":")) 
    # writing the data rows 
    #print(rows)
    csvwriter.writerows(rows)

with open('text_test.csv', 'w') as csvfile:
    csvwriter = csv.writer(csvfile)  
    csvwriter.writerow(fields) 
    rows = []
    for data in sentances:
    # creating a csv writer object 
        
        
    # writing the fields 

        rows.append(data.split(":")) 
    # writing the data rows 
    #print(rows)
    csvwriter.writerows(rows)
    
import pandas as pd
store = []
for element in sentances:
    broken = element.split(':')
    data = ""
    words = broken[0].split(' ')
    for word in words:
        print(word)
        check = True
        for i in range(1,sh.max_row+1):
            keyword = sh.cell(row=i,column=1).value
            if(word == keyword):
                #print("keyword check")
                for j in range(1,sh.max_column+1):
                   
                    #print("one check")
                    one = sh.cell(row=i,column=j).value
                    #print(one)
                    if(one == 1):
                        check = False
                        #print("store check")
                        data= data+" "+str(sh.cell(row=1,column=j).value)
                        break
                break
        if(check == True):
            data = data+" 0"

    store.append(data)
        

# Load the CSV file
df = pd.read_csv('text_train.csv')

# Insert data into column 3

df.insert(2, 'NER', store)

# Write the result back to the CSV file
df.to_csv('your_file.csv', index=False)