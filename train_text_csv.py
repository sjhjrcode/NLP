import pandas as pd
import openpyxl
import numpy as np
import csv

sentances  = ["i want to make coffee:DA"]
sentances.append("where am i:LR")
sentances.append("what can i do here:AR")
sentances.append("prepare meal:DA")
sentances.append("use refrigerator:DA")
sentances.append("use fan:DA")
sentances.append("use oven:DA")
sentances.append("use stove:DA")
sentances.append("i would like to wash sheets:DA")
sentances.append("i would like to watch television:DA")
sentances.append("wash sheets:DA")
sentances.append("wash clothes:DA")
sentances.append("cook dinner:DA")
sentances.append("i would like to cook dinner:DA")
sentances.append("where is the coffee:DA")
sentances.append("rewind video:VC")
sentances.append("forward:VC")

sentances.append("pause video:VC")
sentances.append("skip:VC")
sentances.append("rewind:VC")
sentances.append("pause:VC")
sentances.append("fast forward:VC")
sentances.append("play video:VC")
sentances.append("resume:VC")
sentances.append("resume video:VC")
sentances.append("negative:SN")
sentances.append("positive:SN")
sentances.append("yes:SN")
sentances.append("no:SN")
sentances.append("yes:SN")
sentances.append("no:SN")
sentances.append("shutdown app:SD")
sentances.append("shut:SD")
sentances.append("hello:NA")
#sentances.append("play video")
sentances.append("shut down app:SD")
sentances.append("shut:SD")
#sentances.append("play video")
sentances.append("negative:SN")
sentances.append("positive:SN")
sentances.append("correct:SN")
sentances.append("incorrect:SN")
sentances.append("fake:SN")
sentances.append("false:SN")
sentances.append("nay:SN")
sentances.append("wrong:SN")
sentances.append("yay:SN")

sentances.append("genuine:SN")
sentances.append("true:SN")
sentances.append("right:SN")
sentances.append("correct:SN")
sentances.append("incorrect:SN")
sentances.append("i want to make coffee:DA")
sentances.append("where am i:LR")
sentances.append("start camera:CC")
sentances.append("stop camera:CC")
sentances.append("open camera:CC")
sentances.append("close camera:CC")
sentances.append("start camera:CC")
sentances.append("disengage:CC")
sentances.append("display:CC")

sentances.append("start camera:CC")
sentances.append("stop camera:CC")
sentances.append("open camera:CC")
sentances.append("close camera:CC")

sentances.append("i would like to wash sheets:DA")
sentances.append("i would like to watch television:DA")
sentances.append("wash sheets:DA")
sentances.append("wash clothes:DA")
sentances.append("cook dinner:DA")
sentances.append("i would like to cook dinner:DA")
sentances.append("where is the coffee:DA")
sentances.append("rewind video:VC")
sentances.append("pause video:VC")
sentances.append("skip forward:VC")
sentances.append("shut down:SD")
sentances.append("turn off:SD")
sentances.append("return home:SD")
sentances.append("shutdown:SD")
sentances.append("power off:SD")
sentances.append("shut down system:SD")
sentances.append("system disable:SD")
sentances.append("disengage:SD")
sentances.append("what can I do here:AR")
sentances.append("activites:AR")
sentances.append("what activites are available:AR")
sentances.append("what activites are available at this location:AR")
sentances.append("what can I interactive with:AR")
sentances.append("what can I make:AR")
sentances.append("is there water here:AR")
sentances.append("what is here:AR")
sentances.append("where am I:LR")
sentances.append("where am I currently at:LR")
sentances.append("what is here:AR")
sentances.append("where am is the kitchen:NAV")
sentances.append("where am I currently at:LR")
sentances.append("am I currently in the kitchen:LR")
sentances.append("am I currently in the bedroom:LR")
sentances.append("am I currently in the bathroom:LR")
sentances.append("where am I at:LR")
sentances.append("can you tell me where I am:LR")
sentances.append("i need to know my current position:LR")
sentances.append("location:LR")
sentances.append("position:LR")
sentances.append("where:LR")
sentances.append("activities:AR")
sentances.append("do here:AR")
sentances.append("available:AR")
sentances.append("coffee:DA")
sentances.append("cook:DA")
sentances.append("diswasher:DA")
#sentances.append("society excited by cottage private an it esteems:NA")
#sentances.append("fully begin on by wound an:NA")
#sentances.append("girl rich in do up or both:NA")
#sentances.append("at declared in as rejoiced of together:NA")
#sentances.append("he impression collecting delightful unpleasant by prosperous as on:NA")
#sentances.append("end too talent she object mrs wanted remove giving:NA")
#sentances.append("the text is a corrupted version of the original and therefore does not mean anything in particular:NA")
#sentances.append("It was a scrape that he hardly noticed. Sure, there was a bit of blood but it was minor compared to most of the other cuts and bruises he acquired on his adventures. There was no way he could know that the rock that produced the cut had alien genetic material on it that was now racing through his bloodstream. He felt perfectly normal and continued his adventure with no knowledge of what was about to happen to him.:NA")
#sentances.append("Joe made the sugar cookies Susan decorated them.:NA")
#sentances.append("Being unacquainted with the chief raccoon was harming his prospects for promotion.:NA")
#sentances.append("She was sad to hear that fireflies are facing extinction due to artificial light, habitat loss, and pesticides.:NA")
sentances.append("go to the living room:NAV")
sentances.append("go to kitchen:NAV")
sentances.append("can I go to the living room:NAV")
sentances.append("take me to the sunroom:NAV")
sentances.append("kitchen to bathroom:NAV")
sentances.append("move to garage:NAV")
sentances.append("i want to go to the kitchen:NAV")
sentances.append("how do i get to the bathroom:NAV")
sentances.append("where is the bathroom:NAV")
sentances.append("i am lost take me to the livingroom:NAV")
sentances.append("help me find the kitchen:NAV")
sentances.append("help me to the bedroom:NAV")
df = pd.read_excel("./Keyword Map NLP.xlsx")
Loc = df.loc[df['Loc']==1]
#print(Loc)
workbook = openpyxl.load_workbook('./Keyword Map NLP.xlsx')
sh = workbook.active
'''                 
get= []
print("start")
for i in range(1,sh.max_column+1):
    #print(i)
    search = sh.cell(row=1,column=i).value
    #print(search)
    if(search == 'Loc'):
        for j in range(2,sh.max_row+1):
            one = sh.cell(row=j,column=i).value
            if(one == 1):
                #print(sh.cell(row=j,column=1).value)
                get.append(sh.cell(row=j,column=1).value)
        break
                 
for w in get:
    filer = (f"i want to go to the {w}:NAV")
    sentances.append(filer)

for w in get:
    filer = (f"navigate to {w}:NAV")
    sentances.append(filer)
for w in get:
    filer = (f"take me to {w}:NAV")
    sentances.append(filer)
for w in get:
    filer = (f"where is to {w}:NAV")
    sentances.append(filer)
for w in get:
    filer = (f"can i go to {w}:NAV")
    sentances.append(filer)


for w in get:
    filer = (f"is this the {w}:LR")
    sentances.append(filer)

for w in get:
    filer = (f"am I in the {w}:LR")
    sentances.append(filer)
'''

fields  = ["description","category"]
with open('text_train.csv', 'w') as csvfile:
    csvwriter = csv.writer(csvfile)  
    csvwriter.writerow(fields) 
    rows = []
    for data in sentances:
    # creating a csv writer object 
        
        
    # writing the fields 

        rows.append(data.split(":")) 
    # writing the data rows 
    #print(rows)
    csvwriter.writerows(rows)

with open('text_test.csv', 'w') as csvfile:
    csvwriter = csv.writer(csvfile)  
    csvwriter.writerow(fields) 
    rows = []
    for data in sentances:
    # creating a csv writer object 
        
        
    # writing the fields 

        rows.append(data.split(":")) 
    # writing the data rows 
    #print(rows)
    csvwriter.writerows(rows)