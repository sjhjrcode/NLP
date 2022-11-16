import pandas as pd
import openpyxl
import numpy as np
def sentancebreak(breaking, sh):
    check = True
    broken = breaking.split(' ')
    store = []
    for l in broken:
        check = True
        for i in range(1,sh.max_row+1):
            keyword = sh.cell(row=i,column=1).value
            if(l == keyword):
                print("keyword check")
                for j in range(1,sh.max_column+1):
                   
                    print("one check")
                    one = sh.cell(row=i,column=j).value
                    print(one)
                    if(one == 1):
                        check = False
                        print("store check")
                        store.append(keyword+" 0 0 "+sh.cell(row=1,column=j).value)
                        break
                break
                        

        if(check==True):
            store.append(l+" 0 0 0")

    store.append(". O\n")

    return store


def adding(base, add):
    for k in add:
        base.append(k)
    return base
        
df = pd.read_excel("./Keyword Map NLP.xlsx")
Loc = df.loc[df['Loc']==1]
#print(Loc)
workbook = openpyxl.load_workbook('./Keyword Map NLP.xlsx')
sh = workbook.active
get= []
print("start")
for i in range(1,sh.max_column+1):
    #print(i)
    search = sh.cell(row=1,column=i).value
    #print(search)
    if(search == 'Loc'):
        for j in range(2,sh.max_row+1):
            one = sh.cell(row=j,column=i).value
            if(one == 1):
                #print(sh.cell(row=j,column=1).value)
                get.append(sh.cell(row=j,column=1).value)
        break

sentances = ["-DOCSTART- -X- -X- O\n"]
#file = open("read.txt",'w')

#for i in range(1,sh.max_row+1):
#    keyword = sh.cell(row=i,column=1).value
#    for j in range(1,sh.max_column+1):
#        one = sh.cell(row=i,column=j).value
#        if(one == 1):
#            sentances.append(keyword+" 0 0 "+sh.cell(row=1,column=j).value+"\n")
#            break

for w in get:
    filer = (f"i want to go to the {w}")
    temp = sentancebreak(filer,sh)
    adding(sentances,temp)
for w in get:
    filer = (f"navigate to {w}")
    temp = sentancebreak(filer,sh)
    sentances=adding(sentances,temp)

filer = (f"i want to make coffee")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"where am i")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"what can i do here")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)


sentances=adding(sentances,temp)
filer = (f"prepare meal")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"use refrigerator")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"use fan")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"use oven")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"use stove")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"i would like to wash sheets")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"i would like to watch television")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"wash sheets")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)

filer = (f"wash clothes")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)

filer = (f"cook dinner")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"i would like to cook dinner")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"make food")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"i want to read a book")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)

filer = (f"open camera")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"start camera")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)

filer = (f"what can i do here")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"where is the coffee")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"rewind video")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"pause video")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"skip forward")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"negative")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"positive")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"yes")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"no")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello yes")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello no")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"shutdown app")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"shut")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"play video")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello shut down app")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello shut")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello play video")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello negative")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello positive")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"correct")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"incorrect")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello correct")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello incorrect")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)

with open('eng_custom.train', 'w') as file:
      

    file.writelines("% s\n" % data for data in sentances)



sentances = ["-DOCSTART- -X- -X- O\n"]
#file = open("read.txt",'w')



for w in get:
    filer = (f"hello i want to go to the {w}")
    temp = sentancebreak(filer,sh)
    adding(sentances,temp)
for w in get:
    filer = (f"hello navigate to {w}")
    temp = sentancebreak(filer,sh)
    sentances=adding(sentances,temp)

filer = (f"hello i want to make coffee")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello where am i")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello what can i do here")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello prepare meal")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello use refrigerator")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello start camera")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)

filer = (f"hello correct")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello incorrect")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)


filer = (f"hello play video")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello negative")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello positive")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)


filer = (f"hello yes")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)
filer = (f"hello no")
temp = sentancebreak(filer,sh)
sentances=adding(sentances,temp)


with open('eng_custom.testa', 'w') as file:
      

    file.writelines("% s\n" % data for data in sentances)

